import streamlit as st
from supabase import create_client, Client
import os
import uuid
import random
import string
import requests
import time
import pandas as pd
import io
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# Configuración de página
st.set_page_config(page_title="NotPed - B2B", page_icon="👞", layout="wide")

st.markdown("""
    <style>
        [data-testid="stFileUploader"] small {display: none !important;}
        [data-testid="stFileUploaderDropzoneInstructions"] > div > span {display: none !important;}
        .block-container { padding-top: 2rem; padding-bottom: 2rem; max-width: 95%; }
        div[data-testid="stVerticalBlockBorderWrapper"] {
            padding: 8px 12px !important;
            border-radius: 8px !important;
            box-shadow: 0 1px 2px rgba(0,0,0,0.05) !important;
        }
        .stMarkdown p { margin-bottom: 0px !important; }
        [data-testid="column"] { padding: 0 6px !important; }
        input[type="number"] { text-align: center; font-weight: bold; }
    </style>
""", unsafe_allow_html=True)

@st.cache_resource
def init_connection():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    return create_client(url, key)

supabase: Client = init_connection()

# --- ESTADO DE SESIÓN ---
if "usuario_actual" not in st.session_state: st.session_state.usuario_actual = None
if "rol_actual" not in st.session_state: st.session_state.rol_actual = None
if "marca_actual" not in st.session_state: st.session_state.marca_actual = None
if "form_key" not in st.session_state: st.session_state.form_key = 0 
if "seccion_publica" not in st.session_state: st.session_state.seccion_publica = "inicio"
if "panel_privado" not in st.session_state: st.session_state.panel_privado = "pedidos"
if "remember_email" not in st.session_state: st.session_state.remember_email = ""
if "ndp_activa" not in st.session_state: st.session_state.ndp_activa = None
if "cat_activa" not in st.session_state: st.session_state.cat_activa = None
if "cat_preseleccionada" not in st.session_state: st.session_state.cat_preseleccionada = None

fk = st.session_state.form_key 

# --- RECUPERACIÓN DE SESIÓN POR URL ---
if st.session_state.usuario_actual is None and "uid" in st.query_params:
    try:
        res = supabase.table("usuarios").select("*").eq("id", st.query_params["uid"]).execute()
        if res.data:
            st.session_state.usuario_actual = res.data[0]["email"]
            st.session_state.rol_actual = res.data[0]["rol"]
            st.session_state.marca_actual = res.data[0]["nombre_marca"]
            st.session_state.panel_privado = st.query_params.get("panel", "pedidos")
    except: pass

# --- LINK DIRECTO: AUTO-ASIGNAR CÓDIGO NDP ---
if "codigo" in st.query_params and st.session_state.usuario_actual and st.session_state.rol_actual == "revendedor":
    cod_magico = st.query_params["codigo"]
    res_magico = supabase.table("notas_pedido").select("*").eq("codigo_acceso", cod_magico).execute()
    if res_magico.data:
        ndp_m = res_magico.data[0]
        if ndp_m['estado'] == "Enviado" and not ndp_m['revendedor_email']:
            supabase.table("notas_pedido").update({"revendedor_email": st.session_state.usuario_actual, "estado": "En Proceso"}).eq("id", ndp_m['id']).execute()
            st.session_state.ndp_activa = cod_magico
            st.session_state.panel_privado = "ingresar_ndp"
        elif ndp_m['revendedor_email'] == st.session_state.usuario_actual and ndp_m['estado'] == "En Proceso":
            st.session_state.ndp_activa = cod_magico
            st.session_state.panel_privado = "ingresar_ndp"
    # Quitamos el parametro para evitar bucles
    del st.query_params["codigo"]

# ========================================================
# FUNCIONES AUXILIARES
# ========================================================
def calcular_precio_descuento_cascada(precio_base, desc_str):
    if not desc_str or str(desc_str).strip() == "0": return float(precio_base), False
    try:
        descuentos = [float(d.strip()) for d in str(desc_str).split('+') if d.strip()]
        precio = float(precio_base)
        for d in descuentos: precio *= (1 - d/100)
        return precio, True if descuentos else False
    except: return float(precio_base), False

def mostrar_detalle_pedido(detalle_json, codigo_pedido):
    if not detalle_json: return
    filas = []
    for p_id, data in detalle_json.items():
        fila = {"Artículo": data.get("articulo", ""), "Precio Unit. ($)": data.get("precio_unitario", 0), "Pares Totales": data.get("pares_totales", 0)}
        for talle, cant in data.get("curva_elegida", {}).items(): fila[f"T-{talle}"] = cant
        filas.append(fila)
    if filas:
        df = pd.DataFrame(filas).fillna(0)
        cols_base = ["Artículo", "Precio Unit. ($)", "Pares Totales"]
        cols_talles = sorted([c for c in df.columns if c.startswith("T-")])
        df = df[cols_base + cols_talles]
        for col in cols_talles + ["Pares Totales"]: df[col] = df[col].astype(int)
        st.dataframe(df, use_container_width=True, hide_index=True)
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer: df.to_excel(writer, index=False, sheet_name='Nota de Pedido')
        st.download_button("📊 Descargar Excel", data=output.getvalue(), file_name=f"NdP_{codigo_pedido}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"xls_{codigo_pedido}_{uuid.uuid4().hex[:5]}")

@st.cache_data(show_spinner=False)
def obtener_plantilla_excel():
    df_template = pd.DataFrame({
        "Categoría": ["Zapatillas"], "Artículo": ["Zapa Urban"], "Color": ["Negro"], "Descripción": ["Suela inyectada"],
        "Precio": [15000], "Tipo Curva": ["Numérica Simple"], "Talle Desde": ["35"],
        "Talle Hasta": ["40"], "Cantidades (Ej: 2-2-3-2-2-1)": ["2-2-3-2-2-1"],
        "URL Foto (Opcional)": [""], "URL Video (Opcional)": [""]
    })
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_template.to_excel(writer, index=False, sheet_name='Plantilla')
    return buffer.getvalue()

# ========================================================
# VENTANAS FLOTANTES
# ========================================================
@st.dialog("Renombrar Categoría")
def dialog_editar_categoria(c_id, viejo_nombre):
    nuevo_nombre = st.text_input("Nuevo nombre:", value=viejo_nombre).strip()
    if st.button("Guardar Cambios", type="primary"):
        if nuevo_nombre and nuevo_nombre != viejo_nombre:
            supabase.table("configuraciones_fabrica").update({"nombre": nuevo_nombre}).eq("id", c_id).execute()
            supabase.table("productos").update({"categoria": nuevo_nombre}).eq("proveedor", st.session_state.usuario_actual).eq("categoria", viejo_nombre).execute()
            st.rerun()

@st.dialog("Editar Producto")
def dialog_editar_producto(p):
    n_art = st.text_input("Artículo", value=p['articulo'])
    n_col = st.text_input("Color", value=p.get('color',''))
    n_desc = st.text_input("Descripción", value=p.get('descripcion',''))
    n_precio = st.number_input("Precio ($)", value=float(p['precio']))
    n_vid = st.text_input("URL Video YouTube", value=p.get('video_url', '') or '')
    if st.button("Guardar Cambios", type="primary", use_container_width=True):
        supabase.table("productos").update({
            "articulo": n_art, "color": n_col, "descripcion": n_desc, 
            "precio": n_precio, "video_url": n_vid
        }).eq("id", p['id']).execute()
        st.rerun()

@st.dialog("Subir Foto del Producto")
def dialog_subir_foto(p_id, articulo):
    img_up = st.file_uploader("JPG / PNG (Máx 250KB)", type=["jpg", "png", "jpeg"], key=f"up_{p_id}")
    if img_up:
        if img_up.size > 256000: st.error("❌ La imagen supera los 250KB permitidos.")
        else:
            with st.spinner("Subiendo..."):
                extension = img_up.name.split('.')[-1]
                nombre_archivo = f"{uuid.uuid4()}.{extension}"
                supabase.storage.from_("fotos_productos").upload(nombre_archivo, img_up.getvalue(), {"content-type": img_up.type})
                f_url = supabase.storage.from_("fotos_productos").get_public_url(nombre_archivo)
                supabase.table("productos").update({"foto_url": f_url}).eq("id", p_id).execute()
                st.rerun()

@st.dialog("Vista Ampliada")
def dialog_ampliar_imagen(url, articulo):
    st.image(url, use_container_width=True)

@st.dialog("Compartir Código NdP")
def dialog_compartir_categoria(categoria):
    st.write(f"Compartiendo: **{categoria}**")
    res_clientes = supabase.table("notas_pedido").select("nombre_cliente_referencia").eq("proveedor_email", st.session_state.usuario_actual).execute()
    clientes_unicos = sorted(list(set([c['nombre_cliente_referencia'] for c in res_clientes.data if c.get('nombre_cliente_referencia')])))
    
    opcion_cliente = st.selectbox("Cliente", ["-- Nuevo Cliente --"] + clientes_unicos)
    if opcion_cliente == "-- Nuevo Cliente --":
        cliente_final = st.text_input("Nombre del nuevo cliente:").strip()
        historial_desc = []
    else:
        cliente_final = opcion_cliente
        res_desc = supabase.table("notas_pedido").select("descuento").eq("proveedor_email", st.session_state.usuario_actual).eq("nombre_cliente_referencia", cliente_final).execute()
        historial_desc = sorted(list(set([d['descuento'] for d in res_desc.data if d.get('descuento') and d['descuento'] != "0"])))

    if historial_desc: opcion_desc = st.selectbox(f"Descuentos de {cliente_final}", ["-- Sin Descuento --", "-- Crear Nuevo --"] + historial_desc)
    else: opcion_desc = st.selectbox("Descuento", ["-- Sin Descuento --", "-- Crear Nuevo --"])

    if opcion_desc == "-- Crear Nuevo --": descuento_final = st.text_input("Desc. en Cascada (Ej: 10+5)").strip()
    elif opcion_desc == "-- Sin Descuento --": descuento_final = "0"
    else: descuento_final = opcion_desc
    
    if st.button("Generar Enlace", type="primary"):
        if cliente_final:
            codigo = f"NDP-{uuid.uuid4().hex[:5].upper()}"
            try:
                supabase.table("notas_pedido").insert({
                    "codigo_acceso": codigo, "proveedor_email": st.session_state.usuario_actual, "nombre_cliente_referencia": cliente_final, "categoria_compartida": categoria, "descuento": descuento_final
                }).execute()
                st.success("✅ ¡Código Creado!")
                # Enlace directo
                link_magico = f"https://tunombredeweb.com/?codigo={codigo}"
                st.code(link_magico)
                st.info("Copia el enlace de arriba y envíalo al cliente por WhatsApp. No necesitarán copiar el código, entrarán directo.")
            except Exception as e: st.error(f"Error: {e}")
        else: st.warning("Debes ingresar el nombre del cliente.")

@st.dialog("Duplicar Catálogo Completo")
def dialog_duplicar_categoria(cat_name):
    st.info(f"Vas a crear una copia exacta de todos los productos de **{cat_name}**.")
    nuevo_nombre = st.text_input("Nombre para el nuevo catálogo:", value=f"{cat_name} (Copia)").strip()
    if st.button("Crear Duplicado", type="primary"):
        if nuevo_nombre and nuevo_nombre != cat_name:
            with st.spinner("Clonando catálogo..."):
                prods = supabase.table("productos").select("*").eq("proveedor", st.session_state.usuario_actual).eq("categoria", cat_name).execute().data
                for p in prods:
                    del p['id']
                    if 'created_at' in p: del p['created_at']
                    p['categoria'] = nuevo_nombre
                if prods: supabase.table("productos").insert(prods).execute()
                st.rerun()
        else: st.warning("Ingresa un nombre diferente al original.")

# --- GENERADOR DE PDF ---
@st.cache_data(show_spinner=False)
def generar_pdf_catalogo(productos, marca, titulo_cat):
    try:
        from fpdf import FPDF
        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("helvetica", style="B", size=16)
        pdf.cell(0, 10, f"CATÁLOGO MAYORISTA - {marca}", ln=True, align="C")
        pdf.set_font("helvetica", style="I", size=12)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, f"Categoría: {titulo_cat}", ln=True, align="C")
        pdf.ln(6)
        for p in productos:
            x_start, y_start = pdf.get_x(), pdf.get_y()
            if y_start > 265: pdf.add_page(); y_start = pdf.get_y()
            pdf.set_draw_color(220, 220, 220); pdf.rect(10, y_start, 190, 22)
            if p.get('foto_url'):
                try:
                    resp = requests.get(p['foto_url'], timeout=3)
                    if resp.status_code == 200: pdf.image(io.BytesIO(resp.content), x=12, y=y_start+2, w=22, h=18, keep_aspect_ratio=True)
                except: pass
            pdf.set_text_color(40, 40, 40); pdf.set_xy(38, y_start + 5); pdf.set_font("helvetica", style="B", size=11)
            pdf.cell(70, 6, f"{p['articulo']}  |  {p.get('color','')}", ln=False)
            pdf.set_xy(38, y_start + 11); pdf.set_font("helvetica", size=8); pdf.set_text_color(120, 120, 120)
            desc = str(p.get('descripcion',''))[:45] + "..." if len(str(p.get('descripcion',''))) > 45 else str(p.get('descripcion',''))
            pdf.cell(70, 5, desc, ln=False)
            pdf.set_xy(110, y_start + 5); pdf.set_font("helvetica", size=9); pdf.set_text_color(100, 100, 100)
            if p.get("curva") == "N/A": pdf.cell(40, 6, "Talles: Sin Curva", ln=False)
            else:
                pdf.cell(40, 6, f"Talles: {p['talle_desde']} al {p['talle_hasta']}", ln=False)
                pdf.set_xy(110, y_start + 11); pdf.set_font("courier", style="B", size=9); pdf.set_text_color(23, 162, 184)
                pdf.cell(40, 5, f"Curva: {p.get('curva','')}", ln=False)
            pdf.set_xy(150, y_start + 7); pdf.set_font("helvetica", style="B", size=14); pdf.set_text_color(40, 167, 69)
            pdf.cell(45, 8, f"${p['precio']:,.0f}", ln=False, align="R")
            pdf.set_text_color(0, 0, 0); pdf.set_y(y_start + 24)
        return bytes(pdf.output())
    except: return None

# ========================================================
# FLUX 1: ENTORNO PÚBLICO
# ========================================================
if st.session_state.usuario_actual is None:
    st.query_params.clear() 
    col_logo, col_nav = st.columns([2, 3])
    with col_logo: st.markdown("<h2 style='margin:0;'>👞 NotPed <span style='font-size:14px; color:gray;'>B2B</span></h2>", unsafe_allow_html=True)
    with col_nav:
        sub_col1, sub_col2, sub_col3 = st.columns(3)
        with sub_col1:
            if st.button("🏠 Inicio", use_container_width=True): st.session_state.seccion_publica = "inicio"; st.rerun()
        with sub_col2:
            if st.button("🔐 Iniciar Sesión", use_container_width=True): st.session_state.seccion_publica = "login"; st.rerun()
        with sub_col3:
            if st.button("📝 Registrarse", use_container_width=True): st.session_state.seccion_publica = "registro"; st.rerun()
    st.write("---")

    if st.session_state.seccion_publica == "inicio": st.title("🚀 Conectamos Fábricas con Revendedores")
    elif st.session_state.seccion_publica == "login":
        col_login = st.columns([1, 2, 1])[1]
        with col_login:
            with st.form("login_form"):
                st.subheader("Ingresar")
                u_email = st.text_input("Correo", value=st.session_state.remember_email).strip().lower()
                pwd = st.text_input("Contraseña", type="password").strip()
                recordar = st.checkbox("Recordar mi correo", value=True if st.session_state.remember_email else False)
                if st.form_submit_button("Ingresar"):
                    res = supabase.table("usuarios").select("*").eq("email", u_email).eq("contrasena", pwd).execute()
                    if res.data:
                        st.session_state.usuario_actual = res.data[0]["email"]
                        st.session_state.rol_actual = res.data[0]["rol"]
                        st.session_state.marca_actual = res.data[0]["nombre_marca"]
                        st.session_state.panel_privado = "pedidos" if res.data[0]["rol"] == "proveedor" else "mis_pedidos"
                        st.session_state.remember_email = u_email if recordar else ""
                        st.query_params["uid"] = str(res.data[0]["id"])
                        st.query_params["panel"] = st.session_state.panel_privado
                        st.rerun()
                    else: st.error("❌ Datos incorrectos.")
    elif st.session_state.seccion_publica == "registro":
        col_reg = st.columns([1, 2, 1])[1]
        with col_reg:
            with st.form("reg_form"):
                st.subheader("Crear Cuenta")
                tipo = st.selectbox("Tipo", ["Revendedor", "Fábrica"])
                marca = st.text_input("Marca").strip()
                email = st.text_input("Correo").strip().lower()
                p1 = st.text_input("Clave", type="password")
                p2 = st.text_input("Confirmar", type="password")
                if st.form_submit_button("Registrarse"):
                    if p1 == p2 and marca and email:
                        rol = "proveedor" if tipo == "Fábrica" else "revendedor"
                        chequeo = supabase.table("usuarios").select("id").eq("email", email).execute()
                        if not chequeo.data:
                            supabase.table("usuarios").insert({"email": email, "contrasena": p1, "rol": rol, "nombre_marca": marca}).execute()
                            st.success("✅ Cuenta creada.")
                            time.sleep(1); st.session_state.seccion_publica = "login"; st.rerun()

# ========================================================
# FLUX 2: ENTORNO PRIVADO
# ========================================================
else:
    # ----------------------------------------------------
    # PANEL FÁBRICA
    # ----------------------------------------------------
    if st.session_state.rol_actual == "proveedor":
        with st.sidebar:
            st.markdown(f"### {st.session_state.marca_actual}")
            st.success("🏭 Panel Fábrica")
            st.write("---")
            if st.button("📦 Notas de Pedido", use_container_width=True): 
                st.session_state.panel_privado = "pedidos"
                st.query_params["panel"] = "pedidos"
                st.rerun()
            if st.button("👞 Mis Catálogos", use_container_width=True): 
                st.session_state.cat_activa = None
                st.session_state.panel_privado = "catalogo"
                st.query_params["panel"] = "catalogo"
                st.rerun()
            if st.button("➕ Cargar Calzado", use_container_width=True): 
                st.session_state.panel_privado = "carga"
                st.query_params["panel"] = "carga"
                st.rerun()
            st.write("---")
            
            # --- NUEVA SECCIÓN: PERFIL Y CONTRASEÑA ---
            with st.expander("🔑 Mi Perfil / Contraseña"):
                n_pwd = st.text_input("Nueva contraseña:", type="password", key="pwd_f")
                if st.button("Actualizar", key="btn_pwd_f"):
                    if n_pwd:
                        supabase.table("usuarios").update({"contrasena": n_pwd}).eq("email", st.session_state.usuario_actual).execute()
                        st.success("¡Contraseña actualizada!")
            
            if st.button("🚪 Cerrar Sesión", use_container_width=True):
                st.session_state.clear(); st.query_params.clear(); st.rerun()

        # PEDIDOS
        if st.session_state.panel_privado == "pedidos":
            st.title("📦 Mis Notas de Pedido")
            res_pedidos = supabase.table("notas_pedido").select("*").eq("proveedor_email", st.session_state.usuario_actual).execute().data or []
            
            with st.expander("🔍 Filtros de Búsqueda"):
                f_col1, f_col2, f_col3 = st.columns(3)
                fechas = f_col1.date_input("Rango de Fechas", [])
                clientes_unicos = sorted(list(set([p['nombre_cliente_referencia'] for p in res_pedidos if p.get('nombre_cliente_referencia')])))
                cliente_filtro = f_col2.selectbox("Filtrar por Cliente", ["Todos"] + clientes_unicos)
                estado_filtro = f_col3.selectbox("Estado", ["Todos", "Enviado", "En Proceso", "Finalizado"])
            
            pedidos_filtrados = []
            for p in res_pedidos:
                try: p_date = pd.to_datetime(p.get('created_at')).date()
                except: p_date = None
                pasa_fecha = True
                if len(fechas) == 2 and p_date:
                    if not (fechas[0] <= p_date <= fechas[1]): pasa_fecha = False
                pasa_cliente = (cliente_filtro == "Todos" or p.get('nombre_cliente_referencia') == cliente_filtro)
                pasa_estado = (estado_filtro == "Todos" or p.get('estado') == estado_filtro)
                if pasa_fecha and pasa_cliente and pasa_estado: pedidos_filtrados.append(p)

            finalizados = [p for p in pedidos_filtrados if p['estado'] == 'Finalizado']
            c1, c2, c3 = st.columns(3)
            c1.metric("$ Vendido", f"{sum(p['monto_total'] for p in finalizados):,.0f}")
            c2.metric("Pares Solicitados", f"{sum(p['pares_totales'] for p in finalizados)}")
            c3.metric("NdP Mostradas", f"{len(pedidos_filtrados)}")
            st.write("---")
            
            for ndp in sorted(pedidos_filtrados, key=lambda x: x['id'], reverse=True):
                fecha_str = pd.to_datetime(ndp.get('created_at')).strftime('%d/%m/%Y %H:%M') if ndp.get('created_at') else 'Sin fecha'
                with st.expander(f"[{ndp['estado']}] | Cliente: {ndp['nombre_cliente_referencia']} | Total: ${ndp['monto_total']:,.0f} | 📅 {fecha_str}"):
                    c_desc, c_cod = st.columns([3, 1])
                    c_desc.write(f"**Categoría:** {ndp['categoria_compartida']} | **Descuento:** {ndp.get('descuento', '0')}%")
                    with c_cod: st.code(ndp['codigo_acceso'], language=None)
                    if ndp['estado'] == "Finalizado" and ndp.get('detalle_json'): mostrar_detalle_pedido(ndp['detalle_json'], ndp['codigo_acceso'])
                    elif ndp['estado'] == "En Proceso": st.info("El cliente está armando el pedido en este momento.")

        # CARGA
        elif st.session_state.panel_privado == "carga":
            st.title(f"🏭 Panel Fábrica | Carga de Calzado")
            
            res_conf = supabase.table("configuraciones_fabrica").select("*").eq("proveedor", st.session_state.usuario_actual).execute()
            mis_configs = res_conf.data if res_conf.data else []
            mis_categorias = [c for c in mis_configs if c['tipo'] == 'categoria']
            mis_colores = [c for c in mis_configs if c['tipo'] == 'color']
            
            tipo_carga = st.radio("Modo de Carga", ["Carga Manual", "Carga Masiva (Excel)"], horizontal=True)
            st.write("---")

            if tipo_carga == "Carga Manual":
                c_cat, c_col = st.columns(2)
                with c_cat:
                    if st.session_state.get(f"crear_cat_{fk}", False):
                        col_input, col_btn = st.columns([8, 1])
                        with col_input: cat_final = st.text_input("Categoría", placeholder="Ingresá la nueva categoría...", key=f"txt_cat_{fk}").strip()
                        with col_btn:
                            st.write("&nbsp;")
                            if st.button("🔙", key=f"b_cat_{fk}"): st.session_state[f"crear_cat_{fk}"] = False; st.rerun()
                        es_nueva_cat = True
                    else:
                        opciones_cat = ["-- Elegir --", "➕ Crear Nueva..."] + [c['nombre'] for c in mis_categorias]
                        idx_def_cat = 0
                        cat_pre = st.session_state.get("cat_preseleccionada")
                        if cat_pre in opciones_cat: idx_def_cat = opciones_cat.index(cat_pre)
                        opcion_cat = st.selectbox("Categoría", opciones_cat, index=idx_def_cat, key=f"sel_cat_{fk}")
                        if cat_pre: st.session_state.cat_preseleccionada = None
                        if opcion_cat == "➕ Crear Nueva...": st.session_state[f"crear_cat_{fk}"] = True; st.rerun()
                        cat_final = opcion_cat if opcion_cat != "-- Elegir --" else ""
                        es_nueva_cat = False
                with c_col:
                    if st.session_state.get(f"crear_col_{fk}", False):
                        col_input, col_btn = st.columns([8, 1])
                        with col_input: col_final = st.text_input("Color", placeholder="Ingresá el nuevo color...", key=f"txt_col_{fk}").strip()
                        with col_btn:
                            st.write("&nbsp;")
                            if st.button("🔙", key=f"b_col_{fk}"): st.session_state[f"crear_col_{fk}"] = False; st.rerun()
                        es_nuevo_col = True
                    else:
                        opciones_col = ["-- Elegir --", "➕ Crear Nuevo..."] + [c['nombre'] for c in mis_colores]
                        opcion_col = st.selectbox("Color", opciones_col, key=f"sel_col_{fk}")
                        if opcion_col == "➕ Crear Nuevo...": st.session_state[f"crear_col_{fk}"] = True; st.rerun()
                        col_final = opcion_col if opcion_col != "-- Elegir --" else ""
                        es_nuevo_col = False

                art = st.text_input("Artículo (Ej: Bota 401)", key=f"txt_art_{fk}")
                desc = st.text_input("Descripción detallada", key=f"txt_desc_{fk}")
                
                st.write("---")
                st.markdown("**📏 Configuración de Curva Sugerida**")
                
                cd, ch = st.columns(2)
                t_d_sel = cd.number_input("Talle Desde (Ej: 35)", min_value=1, max_value=60, value=35, key=f"n_d_{fk}")
                t_h_sel = ch.number_input("Talle Hasta (Ej: 40)", min_value=t_d_sel, max_value=60, value=min(t_d_sel+5, 60), key=f"n_h_{fk}")
                
                talles_list_str = [str(i) for i in range(int(t_d_sel), int(t_h_sel)+1)]
                cols_talles = st.columns(len(talles_list_str))
                valores_curva = []
                for i, talle in enumerate(talles_list_str):
                    with cols_talles[i]:
                        val = st.number_input(f"T-{talle}", min_value=0, step=1, value=1, key=f"talle_{i}_{fk}")
                        valores_curva.append(str(val))
                curva_final_str = "-".join(valores_curva)
                
                st.write("---")
                col_p, col_v = st.columns(2)
                with col_p: precio = st.number_input("Precio de Lista ($)", min_value=0.0, key=f"num_precio_{fk}")
                with col_v: video = st.text_input("URL de YouTube (Opcional)", placeholder="Ej: https://youtu.be/...", key=f"txt_vid_{fk}")
                
                foto = st.file_uploader("Subir Foto (Máx 250KB)", type=["jpg", "png", "jpeg"], key=f"file_foto_{fk}")
                
                if st.button("Guardar Producto", type="primary", use_container_width=True):
                    if foto and foto.size > 256000: st.error("❌ La imagen supera los 250KB permitidos.")
                    elif not art or not foto or not cat_final or not col_final: st.warning("⚠️ Faltan datos clave (Artículo, Foto, Categoría o Color).")
                    else:
                        with st.spinner("Guardando..."):
                            try:
                                if es_nueva_cat: supabase.table("configuraciones_fabrica").insert({"proveedor": st.session_state.usuario_actual, "tipo": "categoria", "nombre": cat_final}).execute()
                                if es_nuevo_col: supabase.table("configuraciones_fabrica").insert({"proveedor": st.session_state.usuario_actual, "tipo": "color", "nombre": col_final}).execute()

                                extension = foto.name.split('.')[-1]
                                nombre_archivo = f"{uuid.uuid4()}.{extension}"
                                supabase.storage.from_("fotos_productos").upload(nombre_archivo, foto.getvalue(), {"content-type": foto.type})
                                foto_url = supabase.storage.from_("fotos_productos").get_public_url(nombre_archivo)
                                
                                nuevo_prod = {
                                    "proveedor": st.session_state.usuario_actual, "categoria": cat_final, "articulo": art, "color": col_final, 
                                    "descripcion": desc, "precio": precio, "talle_desde": str(t_d_sel), "talle_hasta": str(t_h_sel),
                                    "curva": curva_final_str, "foto_url": foto_url, "video_url": video
                                }
                                supabase.table("productos").insert(nuevo_prod).execute()
                                
                                st.session_state.form_key += 1
                                st.success("✅ ¡Cargado!")
                                time.sleep(1)
                                st.session_state.panel_privado = "catalogo"
                                st.query_params["panel"] = "catalogo"
                                st.rerun()
                            except Exception as e: st.error(f"Error: {e}")

            elif tipo_carga == "Carga Masiva (Excel)":
                st.info("💡 Descarga la plantilla, llénala y súbela.")
                excel_bytes = obtener_plantilla_excel()
                st.download_button("📥 Descargar Plantilla", data=excel_bytes, file_name="Plantilla.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                st.write("---")
                excel_file = st.file_uploader("Sube tu Excel lleno", type=["xlsx", "xls"])
                if st.button("🚀 Procesar", type="primary") and excel_file is not None:
                    with st.spinner("Procesando..."):
                        try:
                            df = pd.read_excel(excel_file, sheet_name='Plantilla', skiprows=[1]).fillna("")
                            res_conf = supabase.table("configuraciones_fabrica").select("*").eq("proveedor", st.session_state.usuario_actual).execute().data or []
                            nombres_cat_db = [c['nombre'] for c in res_conf if c['tipo'] == 'categoria']
                            nombres_col_db = [c['nombre'] for c in res_conf if c['tipo'] == 'color']
                            
                            nuevas_cats, nuevas_cols = [], []
                            cats_vistas, cols_vistas = set(), set()
                            productos_a_insertar = []
                            
                            for index, row in df.iterrows():
                                cat_xls = str(row.get("Categoría", "")).strip()
                                art_xls = str(row.get("Artículo", "")).strip()
                                col_xls = str(row.get("Color", "")).strip()
                                if not art_xls or not cat_xls: continue 
                                
                                if cat_xls not in nombres_cat_db and cat_xls not in cats_vistas:
                                    nuevas_cats.append({"proveedor": st.session_state.usuario_actual, "tipo": "categoria", "nombre": cat_xls})
                                    cats_vistas.add(cat_xls)
                                if col_xls and col_xls not in nombres_col_db and col_xls not in cols_vistas:
                                    nuevas_cols.append({"proveedor": st.session_state.usuario_actual, "tipo": "color", "nombre": col_xls})
                                    cols_vistas.add(col_xls)
                                    
                                p_val = row.get("Precio", 0)
                                precio_xls = float(p_val) if str(p_val).replace('.','',1).isdigit() else 0.0
                                f_url, v_url = str(row.get("URL Foto (Opcional)", "")).strip(), str(row.get("URL Video (Opcional)", "")).strip()
                                
                                productos_a_insertar.append({
                                    "proveedor": st.session_state.usuario_actual, "categoria": cat_xls, "articulo": art_xls, "color": col_xls, 
                                    "descripcion": str(row.get("Descripción", "")).strip(), "precio": precio_xls,
                                    "talle_desde": str(row.get("Talle Desde", "N/A")).strip(), "talle_hasta": str(row.get("Talle Hasta", "N/A")).strip(),
                                    "curva": str(row.get("Cantidades (Ej: 2-2-3-2-2-1)", "N/A")).strip(), "foto_url": f_url if f_url else None, "video_url": v_url if v_url else None
                                })
                            
                            if nuevas_cats: supabase.table("configuraciones_fabrica").insert(nuevas_cats).execute()
                            if nuevas_cols: supabase.table("configuraciones_fabrica").insert(nuevas_cols).execute()
                            if productos_a_insertar: supabase.table("productos").insert(productos_a_insertar).execute()
                            st.success(f"✅ ¡{len(productos_a_insertar)} productos procesados!"); time.sleep(1.5); st.session_state.panel_privado = "catalogo"; st.rerun()
                        except Exception as e: st.error(f"Error: {e}")

        # CATÁLOGOS
        elif st.session_state.panel_privado == "catalogo":
            if st.session_state.cat_activa is None:
                st.title(f"🏭 Mis Catálogos | {st.session_state.marca_actual}")
                st.write("---")
                res_prod = supabase.table("productos").select("categoria").eq("proveedor", st.session_state.usuario_actual).execute()
                if res_prod.data:
                    categorias_presentes = sorted(list(set([p['categoria'] for p in res_prod.data])))
                    cols = st.columns(3)
                    for idx, cat_name in enumerate(categorias_presentes):
                        with cols[idx % 3]:
                            with st.container(border=True):
                                st.markdown(f"<h3 style='color: #FFB300; text-align:center;'>📁 {cat_name}</h3>", unsafe_allow_html=True)
                                c_abrir, c_comp = st.columns(2)
                                with c_abrir:
                                    if st.button("📂 Abrir", key=f"abr_{idx}", use_container_width=True):
                                        st.session_state.cat_activa = cat_name; st.rerun()
                                with c_comp:
                                    if st.button("🔗 Compartir", key=f"cmp_{idx}", use_container_width=True): dialog_compartir_categoria(cat_name)
                                if st.button("📋 Duplicar Catálogo", key=f"dup_{idx}", use_container_width=True): dialog_duplicar_categoria(cat_name)
                else:
                    st.info("Aún no tienes productos en tu catálogo.")
            else:
                cat_name = st.session_state.cat_activa
                c_back, c_titulo, c_add, c_comp, c_pdf = st.columns([1, 2.5, 1.5, 1.5, 1.5], vertical_alignment="bottom")
                with c_back:
                    if st.button("🔙 Volver"): st.session_state.cat_activa = None; st.rerun()
                with c_titulo:
                    st.markdown(f"<h2 style='color: #FFB300; margin-bottom: 0px;'>📁 {cat_name}</h2>", unsafe_allow_html=True)
                with c_add:
                    if st.button("➕ Añadir Calzado", use_container_width=True, type="primary"):
                        st.session_state.cat_preseleccionada = cat_name
                        st.session_state.panel_privado = "carga"
                        st.query_params["panel"] = "carga"
                        st.rerun()
                with c_comp:
                    if st.button("🔗 Compartir", use_container_width=True): dialog_compartir_categoria(cat_name)
                
                res_prod = supabase.table("productos").select("*").eq("proveedor", st.session_state.usuario_actual).eq("categoria", cat_name).execute()
                prods_cat = sorted(res_prod.data or [], key=lambda x: str(x.get('articulo', '')).lower())
                
                with c_pdf:
                    pdf_data = generar_pdf_catalogo(prods_cat, st.session_state.marca_actual, cat_name)
                    if pdf_data: st.download_button("📥 PDF", data=pdf_data, file_name=f"{cat_name}.pdf", mime="application/pdf", use_container_width=True)
                
                st.write("---")
                for p in prods_cat:
                    with st.container(border=True):
                        c_img, c_lupa, c_vid, c_info, c_talles, c_precio, c_edit, c_del = st.columns([0.8, 0.5, 0.5, 3.5, 2.5, 1.5, 0.5, 0.5], vertical_alignment="center")
                        with c_img:
                            if p.get("foto_url"): st.image(p["foto_url"])
                            else:
                                if st.button("📷", key=f"f_{p['id']}"): dialog_subir_foto(p['id'], p['articulo'])
                        with c_lupa:
                            if p.get("foto_url") and st.button("🔍", key=f"z_{p['id']}"): dialog_ampliar_imagen(p["foto_url"], p['articulo'])
                        with c_vid:
                            if p.get("video_url"): st.link_button("▶️", p["video_url"])
                        with c_info:
                            st.markdown(f"<strong style='font-size:15px;'>{p['articulo']}</strong> | <span style='color:gray;'>{p.get('color','')}</span>", unsafe_allow_html=True)
                        with c_talles:
                            if p.get("curva") == "N/A": st.caption("Sin Curva")
                            else:
                                st.caption(f"Talles: {p.get('talle_desde')} al {p.get('talle_hasta')}")
                                st.code(p.get('curva'))
                        with c_precio:
                            st.markdown(f"<div style='color:#00a650; font-size:22px; font-weight:bold;'>${p['precio']:,.0f}</div>", unsafe_allow_html=True)
                        with c_edit:
                            if st.button("✏️", key=f"e_{p['id']}"): dialog_editar_producto(p)
                        with c_del:
                            if st.button("🗑️", key=f"d_{p['id']}"): supabase.table("productos").delete().eq("id", p['id']).execute(); st.rerun()

    # ----------------------------------------------------
    # PANEL REVENDEDOR
    # ----------------------------------------------------
    elif st.session_state.rol_actual == "revendedor":
        with st.sidebar:
            st.markdown(f"### {st.session_state.marca_actual}")
            st.warning("🛒 Panel Revendedor")
            st.write("---")
            if st.button("🛒 Ingresar código NdP", use_container_width=True): 
                st.session_state.panel_privado = "ingresar_ndp"
                st.query_params["panel"] = "ingresar_ndp"
                st.rerun()
            if st.button("📦 Mis Pedidos", use_container_width=True): 
                st.session_state.panel_privado = "mis_pedidos"
                st.query_params["panel"] = "mis_pedidos"
                st.rerun()
            st.write("---")
            
            # --- NUEVA SECCIÓN: PERFIL Y CONTRASEÑA ---
            with st.expander("🔑 Mi Perfil / Contraseña"):
                n_pwd = st.text_input("Nueva contraseña:", type="password", key="pwd_r")
                if st.button("Actualizar", key="btn_pwd_r"):
                    if n_pwd:
                        supabase.table("usuarios").update({"contrasena": n_pwd}).eq("email", st.session_state.usuario_actual).execute()
                        st.success("¡Contraseña actualizada!")
                        
            if st.button("🚪 Cerrar Sesión", use_container_width=True):
                st.session_state.clear(); st.query_params.clear(); st.rerun()

        if st.session_state.panel_privado == "ingresar_ndp":
            col_in, _ = st.columns([2, 1])
            with col_in:
                cod_input = st.text_input("Ingresar código NdP (Manual):", placeholder="Ej: NDP-A1B2C").strip()
                if st.button("Validar y Cargar Catálogo", type="primary"):
                    if cod_input:
                        res = supabase.table("notas_pedido").select("*").eq("codigo_acceso", cod_input).execute()
                        if res.data:
                            ndp = res.data[0]
                            if ndp['estado'] == "Enviado" and not ndp['revendedor_email']:
                                supabase.table("notas_pedido").update({"revendedor_email": st.session_state.usuario_actual, "estado": "En Proceso"}).eq("id", ndp['id']).execute()
                                st.session_state.ndp_activa = cod_input; st.rerun()
                            elif ndp['revendedor_email'] == st.session_state.usuario_actual and ndp['estado'] == "En Proceso":
                                st.session_state.ndp_activa = cod_input; st.rerun()
                            elif ndp['estado'] == "Finalizado": st.warning("Esta NdP ya fue enviada.")
                            else: st.error("Código no asignado a tu cuenta.")
                        else: st.error("Código no encontrado.")

            if st.session_state.ndp_activa:
                cod = st.session_state.ndp_activa
                ndp_data = supabase.table("notas_pedido").select("*").eq("codigo_acceso", cod).execute().data[0]
                
                desc_str = ndp_data.get('descuento', '0')
                precio_desc_ejemplo, hay_descuento = calcular_precio_descuento_cascada(100, desc_str)
                
                st.write("---")
                if hay_descuento:
                    st.markdown(f"<h3 style='color:#00a650;'>📦 Pedido Activo: {ndp_data['categoria_compartida']} <span style='font-size:16px; background-color:#FFB300; padding:3px 8px; border-radius:5px; color:black;'>Desc: {desc_str}%</span></h3>", unsafe_allow_html=True)
                else:
                    st.markdown(f"<h3 style='color:#00a650;'>📦 Pedido Activo: {ndp_data['categoria_compartida']}</h3>", unsafe_allow_html=True)
                
                prods = supabase.table("productos").select("*").eq("proveedor", ndp_data['proveedor_email'])
                if ndp_data['categoria_compartida'] != "Todo el Catálogo": prods = prods.eq("categoria", ndp_data['categoria_compartida'])
                prods = prods.order("articulo").execute().data
                
                if f"cart_{cod}" not in st.session_state: st.session_state[f"cart_{cod}"] = {}

                for p in prods:
                    precio_final, _ = calcular_precio_descuento_cascada(p['precio'], desc_str)
                    with st.container(border=True):
                        c1, c2, c3, c4 = st.columns([1, 2, 6, 2], vertical_alignment="center")
                        with c1:
                            if p.get('foto_url'): st.image(p['foto_url'])
                        with c2:
                            st.markdown(f"**{p['articulo']}**<br><span style='color:gray; font-size:14px;'>{p.get('color','')}</span>", unsafe_allow_html=True)
                        with c3:
                            st.markdown(f"<span style='font-size:12px; color:#17A2B8;'>Curva Fábrica: {p.get('curva','N/A')}</span>", unsafe_allow_html=True)
                            talles = []
                            if str(p.get('talle_desde','')).isdigit() and str(p.get('talle_hasta','')).isdigit():
                                talles = [str(i) for i in range(int(p['talle_desde']), int(p['talle_hasta'])+1)]
                            else: talles = [f"T-{i+1}" for i in range(len(p.get('curva','').split('-')))] if p.get('curva') != 'N/A' else ["Único"]
                            if not talles: talles = ["Único"]
                            
                            curva_arr = [int(x) for x in p.get('curva','').split('-') if x.isdigit()] if p.get('curva') != 'N/A' else []

                            cols_talles = st.columns(len(talles))
                            detalle_temporal = {}
                            
                            for i, t in enumerate(talles):
                                with cols_talles[i]:
                                    # CARGA PREDETERMINADA DE LA CURVA SUGERIDA
                                    val_sugerido = curva_arr[i] if i < len(curva_arr) else 0
                                    
                                    if f"t_{p['id']}_{t}" not in st.session_state:
                                        if str(p['id']) in st.session_state[f"cart_{cod}"]: 
                                            st.session_state[f"t_{p['id']}_{t}"] = st.session_state[f"cart_{cod}"][str(p['id'])]["curva_elegida"].get(t, val_sugerido)
                                        else: 
                                            st.session_state[f"t_{p['id']}_{t}"] = val_sugerido
                                            
                                    val = st.number_input(t, min_value=0, step=1, key=f"t_{p['id']}_{t}")
                                    if val > 0: detalle_temporal[t] = val

                            item_guardado = st.session_state[f"cart_{cod}"].get(str(p['id']), {}).get("curva_elegida", {})
                            estado_guardado = (detalle_temporal == item_guardado)
                            
                            if estado_guardado and sum(detalle_temporal.values()) > 0: btn_txt, btn_type = "✅ Aplicado", "secondary"
                            elif detalle_temporal == {} and item_guardado == {}: btn_txt, btn_type = "➕ Agregar", "secondary"
                            else: btn_txt, btn_type = "🔴 Aplicar", "primary"
                                
                            if st.button(btn_txt, key=f"btn_add_{p['id']}", type=btn_type):
                                if detalle_temporal:
                                    st.session_state[f"cart_{cod}"][str(p['id'])] = {"articulo": p['articulo'], "precio_unitario": precio_final, "pares_totales": sum(detalle_temporal.values()), "curva_elegida": detalle_temporal}
                                else:
                                    if str(p['id']) in st.session_state[f"cart_{cod}"]: del st.session_state[f"cart_{cod}"][str(p['id'])]
                                st.rerun()

                        with c4:
                            if hay_descuento: st.markdown(f"<span style='text-decoration:line-through; color:gray; font-size:14px;'>${p['precio']:,.0f}</span>", unsafe_allow_html=True)
                            st.markdown(f"<h4 style='color:#00a650; margin:0;'>${precio_final:,.0f}</h4>", unsafe_allow_html=True)
                            pares_temp = sum(detalle_temporal.values())
                            if pares_temp > 0: st.caption(f"Subtotal: ${pares_temp * precio_final:,.0f}")
                                
                st.write("---")
                total_pares = sum(item["pares_totales"] for item in st.session_state[f"cart_{cod}"].values())
                total_plata = sum(item["pares_totales"] * item["precio_unitario"] for item in st.session_state[f"cart_{cod}"].values())
                c_tot1, c_tot2, c_btn = st.columns([3, 3, 4], vertical_alignment="center")
                c_tot1.markdown(f"### Pares Guardados: {total_pares}")
                c_tot2.markdown(f"### Total: ${total_plata:,.0f}")
                
                with c_btn:
                    if st.button("✅ Finalizar Pedido Definitivo", type="primary", use_container_width=True):
                        if not st.session_state[f"cart_{cod}"]: st.error("Aplica al menos un producto.")
                        else:
                            supabase.table("notas_pedido").update({"estado": "Finalizado", "monto_total": total_plata, "pares_totales": total_pares, "detalle_json": st.session_state[f"cart_{cod}"]}).eq("id", ndp_data['id']).execute()
                            st.session_state.ndp_activa = None; st.success("¡Enviado!"); time.sleep(2); st.session_state.panel_privado = "mis_pedidos"; st.rerun()

        elif st.session_state.panel_privado == "mis_pedidos":
            st.title("📦 Mis Pedidos")
            res_pedidos = supabase.table("notas_pedido").select("*").eq("revendedor_email", st.session_state.usuario_actual).execute().data or []
            
            with st.expander("🔍 Filtros de Búsqueda"):
                f_col1, f_col2, f_col3 = st.columns(3)
                fechas = f_col1.date_input("Rango de Fechas", [])
                provs_unicos = sorted(list(set([p['proveedor_email'] for p in res_pedidos if p.get('proveedor_email')])))
                prov_filtro = f_col2.selectbox("Fábrica", ["Todos"] + provs_unicos)
                estado_filtro = f_col3.selectbox("Estado", ["Todos", "Enviado", "En Proceso", "Finalizado"])
            
            pedidos_filtrados = []
            for p in res_pedidos:
                try: p_date = pd.to_datetime(p.get('created_at')).date()
                except: p_date = None
                pasa_fecha = True
                if len(fechas) == 2 and p_date:
                    if not (fechas[0] <= p_date <= fechas[1]): pasa_fecha = False
                if pasa_fecha and (prov_filtro == "Todos" or p.get('proveedor_email') == prov_filtro) and (estado_filtro == "Todos" or p.get('estado') == estado_filtro):
                    pedidos_filtrados.append(p)

            finalizados = [p for p in pedidos_filtrados if p['estado'] == 'Finalizado']
            c1, c2, c3 = st.columns(3)
            c1.metric("$ Invertido", f"{sum(p['monto_total'] for p in finalizados):,.0f}")
            c2.metric("Pares Comprados", f"{sum(p['pares_totales'] for p in finalizados)}")
            c3.metric("NdP Mostradas", f"{len(pedidos_filtrados)}")
            st.write("---")
            
            for ndp in sorted(pedidos_filtrados, key=lambda x: x['id'], reverse=True):
                fecha_str = pd.to_datetime(ndp.get('created_at')).strftime('%d/%m/%Y %H:%M') if ndp.get('created_at') else 'Sin fecha'
                with st.expander(f"[{ndp['estado']}] | Fábrica: {ndp['proveedor_email']} | Total: ${ndp['monto_total']:,.0f} | 📅 {fecha_str}"):
                    c_desc, c_btn_acc = st.columns([3, 1])
                    c_desc.write(f"**Categoría:** {ndp['categoria_compartida']} | **Desc:** {ndp.get('descuento', '0')}%")
                    
                    if ndp['estado'] == "Finalizado" and ndp.get('detalle_json'): 
                        mostrar_detalle_pedido(ndp['detalle_json'], ndp['codigo_acceso'])
                    
                    # --- NUEVO BOTÓN: CONTINUAR PEDIDO DIRECTAMENTE ---
                    elif ndp['estado'] == "En Proceso": 
                        st.info("El pedido está abierto pero sin finalizar.")
                        with c_btn_acc:
                            if st.button("🛒 Abrir / Continuar Pedido", key=f"open_{ndp['id']}", type="primary"):
                                st.session_state.ndp_activa = ndp['codigo_acceso']
                                st.session_state.panel_privado = "ingresar_ndp"
                                st.query_params["panel"] = "ingresar_ndp"
                                st.rerun()
